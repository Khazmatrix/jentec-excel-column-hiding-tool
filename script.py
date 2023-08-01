import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook

def hide_columns(sheet):
    # Define the columns to hide based on sheet name
    if sheet.title in ['Tarlac Chilled', 'Tarlac Dry']:
        columns_to_hide = ['A', 'D', 'E', 'F', 'H']
    elif sheet.title in ['Taytay', 'Taytay Chilled']:
        columns_to_hide = ['C', 'D', 'E', 'H']
    elif sheet.title == 'Raymundo Chilled':
        columns_to_hide = ['A', 'D', 'E', 'F', 'H']
    elif sheet.title == 'Raymundo Frozen':
        columns_to_hide = ['A', 'D', 'E', 'F', 'H', 'I']
    elif sheet.title == 'Mercedes':
        columns_to_hide = ['C', 'E', 'H']
    elif sheet.title == 'Ageing Jennys':
        columns_to_hide = ['A', 'D', 'E', 'F', 'K', 'L']
    elif sheet.title == 'Jennys':
        sheet.column_dimensions.group('A', 'ZZ')  # Hide the entire sheet
        return
    elif sheet.title == 'Iloilo':
        columns_to_hide = ['A', 'D', 'E', 'F', 'H']
    elif sheet.title in ['Davao Dry', 'Davao Chilled', 'Cebu Dry', 'Cebu Chilled']:
        columns_to_hide = ['A', 'D', 'E', 'F', 'H']
    elif sheet.title == 'CDO':
        columns_to_hide = ['A', 'D', 'E', 'G', 'H']
    elif sheet.title == 'Calamba Chilled':
        columns_to_hide = ['C', 'D', 'E', 'F', 'G', 'H', 'J']
    elif sheet.title == 'Calamba Dry':
        columns_to_hide = ['C', 'D', 'E', 'F', 'H']
    elif sheet.title == 'Bacolod':
        columns_to_hide = ['A', 'D', 'E', 'F', 'H']
    else:
        return  # No columns to hide for other sheet names
    
    # Hide the specified columns
    for column in columns_to_hide:
        sheet.column_dimensions[column].hidden = True

# Create the root Tkinter window
root = Tk()
root.withdraw()  # Hide the root window

# Prompt the user to select an Excel file
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Load the workbook
wb = load_workbook(filename=file_path)

# Iterate through all sheets in the workbook
for sheet in wb.sheetnames:
    # Select the sheet
    sheet_obj = wb[sheet]
    
    # Call the function to hide columns
    hide_columns(sheet_obj)

# Save the modified workbook
wb.save(filename=file_path)

# Display completion message
print("Hiding columns from this workbook finished.")
